import csv
import os
import re
import shutil

import pandas as pd
import datetime
from annotation.functions.processing import main

# For .xlsx
import openpyxl
from openpyxl.utils import get_column_letter

# For .xls
import xlrd
from xlutils.copy import copy as xlutils_copy  # To avoid conflict with standard copy

# Label constants
LABEL_LOC_P = "LOC_P"
LABEL_LOC_C = "LOC_C"
LABEL_LOC_B = "LOC_B"
LABEL_HDR = "HDR"
LABEL_DATA = "DATA"
LABEL_BLANK = "BLANK"
LABEL_OTHER = "OTHER"
LABEL_TITLE = "TITLE"
LABEL_NOTE = "NOTE"


def extract_rdo_number(filename):
    """Extract the RDO number (and optional letter) from the filename for sorting."""
    try:
        match = re.search(r'RDO No\. (\d+\w?)\s*-\s*.+\.(?:xls|xlsx)?', filename, re.IGNORECASE)
        return match.group(1) if match else None
    except (ValueError, IndexError) as e:
        print(f"Error processing filename: {filename} - {e}")
        return None


def add_labels_to_excel_file(original_filepath, target_sheet_name, labels_list, new_column_header="Annotation_Label"):
    """
    Adds a list of labels as a new column to an existing Excel sheet,
    preserving original formatting.

    Args:
        original_filepath (str): Path to the Excel file.
        target_sheet_name (str): Name of the sheet to modify.
        labels_list (list): List of labels to add. The length should ideally match
                            the number of data rows in the sheet.
        new_column_header (str): Header for the new column.
    """
    file_ext = os.path.splitext(original_filepath)[1].lower()

    try:
        if file_ext == '.xlsx':
            workbook = openpyxl.load_workbook(original_filepath)
            if target_sheet_name not in workbook.sheetnames:
                print(f"Sheet '{target_sheet_name}' not found in {original_filepath}. Available: {workbook.sheetnames}")
                # Fallback: try to use the active sheet or first sheet if target_sheet_name is problematic
                if workbook.active:
                    sheet = workbook.active
                    print(f"Using active sheet: {sheet.title}")
                elif workbook.sheetnames:
                    sheet = workbook[workbook.sheetnames[0]]
                    print(f"Using first sheet: {sheet.title}")
                else:
                    print("No sheets available in workbook.")
                    return False  # Cannot proceed
            else:
                sheet = workbook[target_sheet_name]

            # Determine the next available column (1-indexed)
            next_col_idx = sheet.max_column + 1 if sheet.max_column > 0 else 1

            print(
                f"Adding labels to {original_filepath}, sheet '{sheet.title}', column {get_column_letter(next_col_idx)}")

            # Add labels to subsequent rows
            for i, label in enumerate(labels_list):
                sheet.cell(row=i + 1, column=next_col_idx, value=label)

            workbook.save(original_filepath)  # Save changes to the same file
            return True

        elif file_ext == '.xls':
            rb = xlrd.open_workbook(original_filepath, formatting_info=True)

            sheet_index = -1
            try:
                sheet_index = rb.sheet_names().index(target_sheet_name)
                rs = rb.sheet_by_index(sheet_index)
            except ValueError:
                print(f"Sheet '{target_sheet_name}' not found in {original_filepath}. Available: {rb.sheet_names()}")
                if rb.sheet_names():
                    rs = rb.sheet_by_index(0)  # Fallback to first sheet
                    sheet_index = 0
                    print(f"Using first sheet: {rs.name}")
                else:
                    print("No sheets available in workbook.")
                    return False

            wb = xlutils_copy(rb)  # Create a writable copy
            ws = wb.get_sheet(sheet_index)  # Get the writable sheet

            # Determine the next available column (0-indexed for xlwt)
            # rs.ncols is the number of columns, so next col index is rs.ncols
            next_col_idx = rs.ncols

            print(f"Adding labels to {original_filepath}, sheet '{rs.name}', column index {next_col_idx}")

            for i, label in enumerate(labels_list):
                ws.write(i, next_col_idx, label)  # xlwt rows are 0-indexed matching labels_list index

            wb.save(original_filepath)  # Save changes
            return True

        else:
            print(f"Unsupported file format: {file_ext} for {original_filepath}")
            return False
    except Exception as e:
        print(f"Error adding labels to Excel file {original_filepath}: {e}")
        return False


def xls_to_df(filename, base_dir="data/", full_path=None):
    """Convert Excel file to DataFrame and return with sheet name"""
    filepath = full_path if full_path else os.path.join(base_dir, filename)

    try:
        # Determine engine based on file extension
        if filename.lower().endswith('.xls'):
            excel_file = pd.ExcelFile(filepath, engine='xlrd')
        elif filename.lower().endswith('.xlsx'):
            excel_file = pd.ExcelFile(filepath, engine='openpyxl')
        else:
            excel_file = pd.ExcelFile(filepath)  # Use pandas default engine detection

        # Find and sort sheets matching "Sheet<number>" pattern
        sheet_names = excel_file.sheet_names
        relevant_sheets = sorted(
            [name for name in sheet_names if name.strip().lower().startswith('sheet')],
            key=lambda name: int(re.search(r'\d+', name).group()) if re.search(r'\d+', name) else -1
        )

        if relevant_sheets:
            selected_sheet = relevant_sheets[-1]  # Select the last sheet
            return pd.read_excel(excel_file, sheet_name=selected_sheet, header=None), selected_sheet
        else:
            print(f"No matching sheets found in {filename}")
            return None, None

    except Exception as e:
        print(f"Error processing file {filename}: {e}")
        return None, None


def process_excel_files(input_dir="../data/", base_output_dir="output_files", use_alternate=False, debug=False,
                        initial_version=1):
    """Process Excel files and save results with versioned output"""
    global_annotations_list = []

    # Create versioned output directory
    today = datetime.datetime.now()
    date_str = today.strftime("%d_%m")

    version = initial_version
    while True:
        output_dir_candidate = os.path.join(base_output_dir, f"output_v{version}_{date_str}")
        if not os.path.exists(base_output_dir) or not os.listdir(base_output_dir):  # If base_output_dir is new or empty
            output_dir = output_dir_candidate
            break
        version_pattern = f"output_v{version}_"
        version_exists = any(
            d.startswith(version_pattern) and os.path.isdir(os.path.join(base_output_dir, d))
            for d in os.listdir(base_output_dir)  # list contents of base_output_dir
        )
        if not version_exists:
            output_dir = output_dir_candidate
            break
        version += 1

    print(f"Creating output directory: {output_dir}")
    os.makedirs(output_dir, exist_ok=True)

    # Get files to process
    files_to_process = [f for f in os.listdir(input_dir) if os.path.isfile(os.path.join(input_dir, f)) and
                        f.lower().endswith(('.xls', '.xlsx'))]

    processed_count = 0
    skipped_count = 0

    # Process each file
    for file in files_to_process:
        print(f'Processing {file}')
        original_file_path = os.path.join(input_dir, file)
        df, sheet_name = xls_to_df(file, base_dir=input_dir)

        if df is not None and sheet_name is not None:
            rdo_number = extract_rdo_number(file)
            rdo_id = f"RDO_{rdo_number}" if rdo_number else "RDO_UNKNOWN"
            clean_sheet_name = re.sub(r'[^a-zA-Z0-9]', '_', sheet_name) if sheet_name else "Unknown"

            current_sheet_annotations_cache = {}

            # Process the data
            structured_data = main(
                df,
                filename_for_ann=file,
                sheetname_for_ann=sheet_name,
                annotations_cache=current_sheet_annotations_cache,
                debug=debug,
                debug_location=debug,
                debug_header=debug
            )

            if current_sheet_annotations_cache:
                labels = [entry["label"] for entry in current_sheet_annotations_cache.values()]

                for row_idx in sorted(current_sheet_annotations_cache.keys()):
                    global_annotations_list.append(current_sheet_annotations_cache[row_idx])

            # Save processed data
            output_path = os.path.join(output_dir, f"{rdo_id}_{clean_sheet_name}.csv")
            structured_data.to_csv(output_path, index=False)

            # save edited raw
            labelled_output_path = os.path.join(output_dir, "labels")
            os.makedirs(labelled_output_path, exist_ok=True)

            output_excel_filename = f"{os.path.splitext(file)[0]}_with_labels{os.path.splitext(file)[1]}"
            output_path_excel = os.path.join(labelled_output_path, output_excel_filename)
            shutil.copy2(original_file_path, output_path_excel)  # Copy original to output dir
            print(f"Copied '{original_file_path}' to '{output_path_excel}'")

            # 5. Add labels to the *copied* Excel file, preserving formatting
            success = add_labels_to_excel_file(
                output_path_excel,  # Modify the copy
                sheet_name,
                labels
            )

            print(f'Saved as: {output_path}')
            processed_count += 1
        else:
            print(f"Could not process {file}. Skipping...")
            skipped_count += 1

    # Process annotations
    if global_annotations_list:
        final_annotations_to_save = sorted(
            global_annotations_list,
            key=lambda x: (x["filename"], x["sheetname"], x["row_index"])
        )

        ann_output_path = os.path.join(output_dir, "annotations.csv")
        print(f"\nWriting {len(final_annotations_to_save)} annotations to {ann_output_path}")

        try:
            annotation_df = pd.DataFrame(final_annotations_to_save)
            # Ensure desired column order
            cols_order = ["filename", "sheetname", "row_index", "label", "raw_cells_json"]
            cols_to_use = [col for col in cols_order if col in annotation_df.columns]
            if cols_to_use:
                annotation_df[cols_to_use].to_csv(ann_output_path, index=False, quoting=csv.QUOTE_ALL)
                print(f"Annotations saved successfully to {ann_output_path}")
            else:
                print("Annotation DataFrame was empty or had unexpected columns.")
        except Exception as e:
            print(f"Error writing annotations CSV: {e}")
    else:
        print("No annotations were generated.")

    print(f"\nProcessing complete: {processed_count} files processed, {skipped_count} files skipped.")
    print(f"Output directory: {output_dir}")

    return output_dir


def compare_excel_files(file1, file2, num_rows=3, verbose=False, full_diff=False, output_path=None):
    """
    Compare two Excel files or pandas DataFrames and identify rows that are different.
    When a difference is found, print the row index and the next few rows from each file.

    Args:
        file1 (str or pandas.DataFrame): Path to the first Excel file or a pandas DataFrame
        file2 (str or pandas.DataFrame): Path to the second Excel file or a pandas DataFrame
        num_rows (int, optional): Number of rows to display when a difference is found. Default is 3.
        verbose (bool, optional): Whether to print detailed information. Default is False.
        full_diff (bool, optional): Whether to continue checking after first difference. Default is False.
        output_path (str, optional): Path to save the comparison results. Default is None.
    """
    import pandas as pd

    # Function to clean values if needed
    def clean_value(value):
        # Add your cleaning logic here if needed
        return value

    # Read the Excel files or use provided DataFrames
    try:
        if isinstance(file1, str):
            print(f"Reading file: {file1}")
            df1 = pd.read_excel(file1)
        else:
            print("Using provided DataFrame for file 1")
            df1 = file1.copy()

        if isinstance(file2, str):
            print(f"Reading file: {file2}")
            df2 = pd.read_excel(file2)
        else:
            print("Using provided DataFrame for file 2")
            df2 = file2.copy()

        # Get file names or identifiers for reporting
        file1_name = file1 if isinstance(file1, str) else "DataFrame 1"
        file2_name = file2 if isinstance(file2, str) else "DataFrame 2"

        print(f"Comparing {file1_name} and {file2_name}...")

        # Apply any necessary transformations (customize as needed)
        if "ZV/SQM" in df2.columns:
            df2["ZV/SQM"] = df2["ZV/SQM"].map(clean_value)

    except Exception as e:
        print(f"Error reading data: {e}")
        return

    # Check if DataFrames have the same shape
    if df1.shape != df2.shape:
        print(f"Data sources have different dimensions:")
        print(f"Source 1: {df1.shape[0]} rows, {df1.shape[1]} columns")
        print(f"Source 2: {df2.shape[0]} rows, {df2.shape[1]} columns")

        # Continue with comparison of overlapping rows
        min_rows = min(df1.shape[0], df2.shape[0])
        print(f"Comparing only the first {min_rows} rows...")
        df1 = df1.iloc[:min_rows, :]
        df2 = df2.iloc[:min_rows, :]

    # Initialize list to store comparison results
    comparison_results = []

    # Initialize counters for different types of non-critical differences
    difference_counters = {
        "case_differences": 0,
        "missing_value_differences": 0,
        "numeric_equality_differences": 0,
        "total_checked_cells": 0,
        "real_differences": 0,
        "rows_with_differences": 0,
        "rows_with_real_differences": 0
    }

    # Compare the DataFrames row by row
    differences_found = False
    for idx in range(len(df1)):
        if not df1.iloc[idx].equals(df2.iloc[idx]):
            differences_found = True
            difference_counters["rows_with_differences"] += 1

            # Don't print row difference yet - wait until we confirm it's a real difference

            # Get the next few rows or remaining rows if less than requested
            rows_to_display = min(num_rows, len(df1) - idx)

            if verbose:
                # Print the difference
                print(f"\nSource 1 ({file1_name}) rows {idx} to {idx + rows_to_display - 1}:")
                print(df1.iloc[idx:idx + rows_to_display].to_string())

                print(f"\nSource 2 ({file2_name}) rows {idx} to {idx + rows_to_display - 1}:")
                print(df2.iloc[idx:idx + rows_to_display].to_string())

            # Highlight specific differences in the first different row
            different_columns = []
            for col in df1.columns:
                if col in df2.columns and df1.iloc[idx][col] != df2.iloc[idx][col]:
                    different_columns.append(col)
                    difference_counters["total_checked_cells"] += 1

            # Process differences to identify "real" differences
            real_differences = []

            # Check each column with differences
            for col in different_columns:
                # Check for case differences in string columns
                if isinstance(df1.iloc[idx][col], str) and isinstance(df2.iloc[idx][col], str):
                    if df1.iloc[idx][col].lower() == df2.iloc[idx][col].lower():
                        difference_counters["case_differences"] += 1
                        if verbose:
                            print(f"Column '{col}': Only case difference")
                        continue  # This column has only case difference

                # Check for all kinds of missing value differences
                elif (pd.isna(df1.iloc[idx][col]) or
                      df1.iloc[idx][col] == '' or
                      df1.iloc[idx][col] is None or
                      (isinstance(df1.iloc[idx][col], str) and df1.iloc[idx][col].lower() in ['nan', 'none', 'null',
                                                                                              'na'])) and \
                        (pd.isna(df2.iloc[idx][col]) or
                         df2.iloc[idx][col] == '' or
                         df2.iloc[idx][col] is None or
                         (isinstance(df2.iloc[idx][col], str) and df2.iloc[idx][col].lower() in ['nan', 'none', 'null',
                                                                                                 'na'])):
                    difference_counters["missing_value_differences"] += 1
                    if verbose:
                        print(f"Column '{col}': Both values are missing (NaN/None/empty/string representations)")
                    continue  # This column has only missing value difference

                # Check for numeric equality, with special handling for ZV/SQM column
                elif col == "ZV/SQM" or (
                        isinstance(df1.iloc[idx][col], (int, float)) and isinstance(df2.iloc[idx][col], (int, float))):
                    try:
                        val1 = float(df1.iloc[idx][col]) if not pd.isna(df1.iloc[idx][col]) else None
                        val2 = float(df2.iloc[idx][col]) if not pd.isna(df2.iloc[idx][col]) else None
                        if val1 == val2:
                            difference_counters["numeric_equality_differences"] += 1
                            if verbose:
                                print(
                                    f"Column '{col}': Values are numerically equal ({df1.iloc[idx][col]} vs {df2.iloc[idx][col]})")
                            continue  # Values are numerically equal
                    except (ValueError, TypeError):
                        # If conversion fails, fall through to regular comparison
                        pass

                # If we get here, this is a real difference
                real_differences.append(col)
                difference_counters["real_differences"] += 1

            # If no real differences remain, skip this row
            if not real_differences:
                # Don't print anything for non-real differences
                if verbose:
                    print(
                        f"Row {idx}: Only case differences, equivalent missing values, or numerically equal values found")
                continue  # Skip this difference and continue checking other rows
            else:
                # Now that we know it's a real difference, create the comparison info
                diff_info = {
                    "row_index": idx,
                    "file1_rows": df1.iloc[idx:idx + rows_to_display].to_dict('records'),
                    "file2_rows": df2.iloc[idx:idx + rows_to_display].to_dict('records')
                }
                comparison_results.append(diff_info)

                # Replace different_columns with real_differences for reporting
                different_columns = real_differences
                difference_counters["rows_with_real_differences"] += 1

                # Print the difference header only for real differences
                print(f"\n===== Difference found at row {idx} =====")

                # Print the specific differences
                print("\nSpecific differences in row", idx, ":")
                for col in different_columns:
                    print(f"Column '{col}':")
                    print(f"  Source 1: {df1.iloc[idx][col]}")
                    print(f"  Source 2: {df2.iloc[idx][col]}")

            if not full_diff:
                break

    if not differences_found:
        print("No differences found. The data sources are identical.")
    else:
        # Print summary of differences
        print("\n===== Summary of Differences =====")
        print(f"Total rows checked: {len(df1)}")
        print(f"Rows with any differences: {difference_counters['rows_with_differences']}")
        print(f"Rows with real differences: {difference_counters['rows_with_real_differences']}")
        print(f"Total cells with differences checked: {difference_counters['total_checked_cells']}")
        print(f"Case differences (ignored): {difference_counters['case_differences']}")
        print(f"Missing value differences (ignored): {difference_counters['missing_value_differences']}")
        print(f"Numeric equality differences (ignored): {difference_counters['numeric_equality_differences']}")
        print(f"Real differences (reported): {difference_counters['real_differences']}")

    # Save results to output file if specified
    if output_path and comparison_results:
        with open(output_path, 'w') as f:
            f.write(f"Comparison between {file1_name} and {file2_name}\n\n")
            for diff in comparison_results:
                idx = diff["row_index"]
                f.write(f"Difference at row {idx}:\n")

                f.write(f"\nSource 1 rows {idx} to {idx + len(diff['file1_rows']) - 1}:\n")
                f.write(pd.DataFrame(diff["file1_rows"]).to_string())

                f.write(f"\nSource 2 rows {idx} to {idx + len(diff['file2_rows']) - 1}:\n")
                f.write(pd.DataFrame(diff["file2_rows"]).to_string())

                f.write("\n" + "=" * 50 + "\n")

            # Write summary information to file as well
            f.write("\n===== Summary of Differences =====\n")
            f.write(f"Total rows checked: {len(df1)}\n")
            f.write(f"Rows with any differences: {difference_counters['rows_with_differences']}\n")
            f.write(f"Rows with real differences: {difference_counters['rows_with_real_differences']}\n")
            f.write(f"Total cells with differences checked: {difference_counters['total_checked_cells']}\n")
            f.write(f"Case differences (ignored): {difference_counters['case_differences']}\n")
            f.write(f"Missing value differences (ignored): {difference_counters['missing_value_differences']}\n")
            f.write(f"Numeric equality differences (ignored): {difference_counters['numeric_equality_differences']}\n")
            f.write(f"Real differences (reported): {difference_counters['real_differences']}\n")

        print(f"\nComparison results saved to {output_path}")

    return comparison_results
